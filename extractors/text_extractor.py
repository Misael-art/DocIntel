"""
DocIntel - Text Extractor Module (Stage 3)
Extracts metadata and text from various document formats (PDF, DOCX, XLSX, TXT).
Respects exclusion rules for development directories.
"""
import os
import pypdf
import docx
import openpyxl
from .hasher import should_exclude

def extract_text_from_pdf(pdf_path):
    """Extracts text from a PDF file."""
    try:
        reader = pypdf.PdfReader(pdf_path)
        text = ""
        for page in reader.pages:
            text += page.extract_text() + "\n"
        return text.strip()
    except Exception:
        return ""

def extract_text_from_docx(docx_path):
    """Extracts text from a DOCX file."""
    try:
        doc = docx.Document(docx_path)
        text = "\n".join([para.text for para in doc.paragraphs])
        return text.strip()
    except Exception:
        return ""

def extract_text_from_xlsx(xlsx_path):
    """Extracts text/metadata from an XLSX file."""
    try:
        wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
        sheet_names = wb.sheetnames
        content = f"Sheets: {', '.join(sheet_names)}\n"
        # Extract a small sample from the first sheet
        sheet = wb[sheet_names[0]]
        for row in sheet.iter_rows(max_row=20, max_col=10, values_only=True):
            content += " ".join([str(cell) for cell in row if cell is not None]) + "\n"
        return content.strip()
    except Exception:
        return ""

def extract_content(file_path):
    """Orchestrates text extraction based on file extension."""
    if should_exclude(file_path):
        return None
        
    _, ext = os.path.splitext(file_path)
    ext = ext.lower()
    
    if ext == '.pdf':
        return extract_text_from_pdf(file_path)
    elif ext == '.docx' or ext == '.doc':
        return extract_text_from_docx(file_path)
    elif ext == '.xlsx' or ext == '.xls':
        return extract_text_from_xlsx(file_path)
    elif ext == '.txt' or ext == '.md':
        try:
            with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
                return f.read(10000) # Read first 10k chars
        except Exception:
            return ""
    return ""

if __name__ == "__main__":
    # Test
    test_file = r"F:\DocIntel\walkthrough.md"
    content = extract_content(test_file)
    if content:
        print(f"Extracted content length: {len(content)}")
    else:
        print("Extraction skipped or failed.")
